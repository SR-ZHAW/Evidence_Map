#!/usr/bin/env python3
"""
RIS Folder Screener
===================
Rules:
  - Title must contain at least one term from STRING_A (technology)
  - Title OR abstract must contain at least one term from STRING_B (geography)
  - Both conditions must be met for a YES

Output:
  - Excel file: all records, YES/NO labelled, with links and abstracts
  - RIS file:   YES records only, abstracts removed (PoP abstracts stripped)

SETUP (run once in PyCharm terminal):
    pip install openpyxl
"""

import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# <<<  CONFIGURE HERE  >>>
# ══════════════════════════════════════════════════════════════════

# Folder containing your .ris files
RIS_FOLDER = ""

# Where to save the two output files (e.g. your Desktop)
OUTPUT_FOLDER = ""

# Output file names (you can rename these)
OUTPUT_EXCEL_NAME = ""
OUTPUT_RIS_NAME = ""

# ── STRING A — Technology terms ───────────────────────────────────
# Must appear in the TITLE
STRING_A = """
aquaponic* OR hydroponic* OR aeroponic* OR bioponic* OR soilless* OR
nutrient film technique OR media bed OR ebb and flow system OR
deep water culture OR deep flow technique OR
recirculating aquaculture OR biofloc* OR flocponic* OR raceway* OR
pond culture OR pond aquaculture OR pond fish* OR fish pond* OR
land-based aquaculture OR landlocked aquaculture OR
integrated aquaculture OR
Integrated aqua-agriculture OR Integrated agri-aquaculture OR
Integrated agriculture-aquaculture OR Integrated aquaculture-agriculture OR Aquaculture OR Pond OR Fish OR RAS OR DFT OR drip irrigation OR Plant OR Greens
"""

# ── STRING B — Geography terms ────────────────────────────────────
# Can appear in TITLE or ABSTRACT
STRING_B = """
SSA OR "Sub-Saharan Africa" OR "Sub Saharan Africa" OR Saharan OR Angola OR Benin OR Botswana OR "Burkina Faso" OR Burundi OR Cameroon OR "Cape Verde" OR "Central African Republic" OR Chad OR Comoros OR Congo OR "Côte d'Ivoire" OR "Ivory Coast" OR "Democratic Republic of Congo" OR Djibouti OR "Equatorial Guinea" OR Eritrea OR Eswatini OR Ethiopia OR Gabon OR Gambia OR Ghana OR Guinea OR "Guinea-Bissau" OR Kenya OR Lesotho OR Liberia OR Madagascar OR Malawi OR Mali OR Mauritania OR Mauritius OR Mozambique OR Namibia OR Niger OR Nigeria OR Rwanda OR "São Tomé and Príncipe" OR Senegal OR Seychelles OR "Sierra Leone" OR Somalia OR "South Africa" OR "South Sudan" OR Sudan OR Tanzania OR Togo OR Uganda OR Zambia OR Zimbabwe OR "West Africa" OR "East Africa" OR "Southern Africa" OR "Central Africa"
"""
#addes wild card terms: Aquaculture OR Pond OR Fish OR RAS OR DFT OR drip irrigation OR Plant OR Greens AND Africa

# ══════════════════════════════════════════════════════════════════
# END CONFIGURATION
# ══════════════════════════════════════════════════════════════════


# ------------------------------------------------------------------
# 1.  RIS PARSING
# ------------------------------------------------------------------
TAG_RE = re.compile(r"^([A-Z][A-Z0-9])  - (.*)$")

def parse_ris_file(filepath):
    records = []
    current = {}
    current_field = None
    with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
        for raw_line in f:
            line = raw_line.rstrip("\n").rstrip("\r")
            if not line.strip():
                continue
            match = TAG_RE.match(line)
            if match:
                tag, value = match.groups()
                current_field = tag
                if tag == "ER":
                    if current:
                        records.append(current)
                    current = {}
                    current_field = None
                elif tag == "AU":
                    current.setdefault("AU", []).append(value)
                else:
                    current[tag] = value
            else:
                if current_field and current_field in current:
                    if current_field == "AU":
                        if current["AU"]:
                            current["AU"][-1] += " " + line.strip()
                    else:
                        current[current_field] += " " + line.strip()
    if current:
        records.append(current)
    return records


def parse_ris_folder(folder):
    all_records = []
    ris_files = [f for f in os.listdir(folder) if f.lower().endswith(".ris")]
    if not ris_files:
        print(f"No .ris files found in: {folder}")
        return []
    print(f"Found {len(ris_files)} RIS file(s): {', '.join(ris_files)}")
    for fname in sorted(ris_files):
        path = os.path.join(folder, fname)
        records = parse_ris_file(path)
        for r in records:
            r["_source_file"] = fname
        all_records.extend(records)
        print(f"  {fname}: {len(records)} records")
    return all_records


# ------------------------------------------------------------------
# 2.  HELPERS
# ------------------------------------------------------------------
def get_title(r):
    return r.get("TI") or r.get("T1") or ""

def get_authors(r):
    return "; ".join(r.get("AU", []))

def get_access_link(r):
    url = r.get("UR", "").strip()
    if url:
        return url.split(";")[0].strip()
    doi = r.get("DO", "").strip()
    if doi:
        return doi if doi.lower().startswith("http") else f"https://doi.org/{doi}"
    return ""


# ------------------------------------------------------------------
# 3.  KEYWORD SCREENING
#     STRING_A: title only
#     STRING_B: title OR abstract
# ------------------------------------------------------------------
def parse_terms(raw):
    terms = re.split(r'\s+OR\s+', raw, flags=re.IGNORECASE)
    return [t.strip().strip('"').strip("'") for t in terms if t.strip()]

def compile_term(term):
    has_wildcard = term.endswith("*")
    base = term.rstrip("*")
    escaped = re.escape(base)
    pattern = (escaped + r"\w*") if has_wildcard else escaped
    return re.compile(r"\b" + pattern + r"\b", re.IGNORECASE)

def screen_records(records, string_a_raw, string_b_raw):
    compiled_a = [(t, compile_term(t)) for t in parse_terms(string_a_raw)]
    compiled_b = [(t, compile_term(t)) for t in parse_terms(string_b_raw)]

    n_yes = n_a_only = n_b_only = n_neither = 0

    for r in records:
        title         = get_title(r)
        title_and_ab  = f"{title} {r.get('AB', '')}"

        # STRING_A: title only
        matched_a = [t for t, p in compiled_a if p.search(title)]
        # STRING_B: title OR abstract
        matched_b = [t for t, p in compiled_b if p.search(title_and_ab)]

        hits_a = len(matched_a) > 0
        hits_b = len(matched_b) > 0

        if hits_a and hits_b:
            r["_match"] = "YES"
            n_yes += 1
        else:
            r["_match"] = "NO"
            if hits_a and not hits_b:
                n_a_only += 1
            elif hits_b and not hits_a:
                n_b_only += 1
            else:
                n_neither += 1

        r["_matched_a"] = matched_a
        r["_matched_b"] = matched_b

    print(f"\nScreening complete:")
    print(f"  YES — technology in title + geography in title/abstract : {n_yes}")
    print(f"  NO  — technology in title only, no geography found      : {n_a_only}")
    print(f"  NO  — geography found, but no technology term in title  : {n_b_only}")
    print(f"  NO  — matched neither string                            : {n_neither}")
    print(f"  Total                                                   : {len(records)}")
    return records


# ------------------------------------------------------------------
# 4.  EXCEL EXPORT  (all records, includes PoP abstract for context)
# ------------------------------------------------------------------
def write_excel(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    headers = [
        "Source File", "Title", "Abstract (PoP)",
        "Authors", "Year", "Journal",
        "MATCH",
        "Matched — Technology terms (A, title)",
        "Matched — Geography terms (B, title/abstract)",
        "Access Link"
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    wrap     = Alignment(wrap_text=True, vertical="top")

    for row_idx, r in enumerate(records, start=2):
        match = r.get("_match", "NO")
        values = [
            r.get("_source_file", ""),
            get_title(r),
            r.get("AB", ""),
            get_authors(r),
            r.get("PY", ""),
            r.get("JF", ""),
            match,
            "; ".join(r.get("_matched_a", [])),
            "; ".join(r.get("_matched_b", [])),
            get_access_link(r)
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = wrap

        mc = ws.cell(row=row_idx, column=7)
        mc.fill      = yes_fill if match == "YES" else no_fill
        mc.font      = Font(name="Calibri", size=10, bold=True)
        mc.alignment = Alignment(horizontal="center", vertical="top")

        link = get_access_link(r)
        lc = ws.cell(row=row_idx, column=10)
        if link:
            lc.value     = "Open link"
            lc.hyperlink = link
            lc.font      = Font(name="Calibri", size=10, color="0563C1", underline="single")

    widths = [18, 42, 55, 28, 8, 25, 8, 35, 35, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"
    wb.save(output_path)
    print(f"\n  Excel saved -> {output_path}")


# ------------------------------------------------------------------
# 5.  RIS EXPORT  (YES records only, abstract removed)
# ------------------------------------------------------------------
def write_ris(records, output_path):
    yes_records = [r for r in records if r.get("_match") == "YES"]
    with open(output_path, "w", encoding="utf-8") as f:
        for r in yes_records:
            f.write(f"TY  - {r.get('TY', 'JOUR')}\n")
            title = get_title(r)
            if title:
                f.write(f"TI  - {title}\n")
            for author in r.get("AU", []):
                f.write(f"AU  - {author}\n")
            for tag in ("PY", "JF", "DO", "UR", "PB"):
                if r.get(tag):
                    f.write(f"{tag}  - {r[tag]}\n")
            # AB field intentionally omitted (PoP abstracts are incomplete)
            a_note = "; ".join(r.get("_matched_a", []))
            b_note = "; ".join(r.get("_matched_b", []))
            note = (f"SCREENED YES | "
                    f"Technology terms in title: {a_note} | "
                    f"Geography terms in title/abstract: {b_note}")
            existing = r.get("N1", "")
            f.write(f"N1  - {existing + ' | ' + note if existing else note}\n")
            f.write("ER  - \n\n")
    print(f"  RIS  saved  -> {output_path}  ({len(yes_records)} included records)")


# ------------------------------------------------------------------
# 6.  MAIN
# ------------------------------------------------------------------
def main():
    print("=" * 60)
    print("  RIS FOLDER SCREENER")
    print("  STRING_A in title  AND  STRING_B in title or abstract")
    print("=" * 60)

    if not os.path.isdir(RIS_FOLDER):
        print(f"\n  RIS_FOLDER not found: {RIS_FOLDER}")
        print("  Edit the RIS_FOLDER variable at the top of this script.")
        return

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    print("\n[ STEP 1 ]  Loading RIS files ...")
    records = parse_ris_folder(RIS_FOLDER)
    if not records:
        return
    print(f"\nTotal records loaded: {len(records)}")

    print("\n[ STEP 2 ]  Screening records ...")
    records = screen_records(records, STRING_A, STRING_B)

    print("\n[ STEP 3 ]  Writing output files ...")
    write_excel(records, os.path.join(OUTPUT_FOLDER, OUTPUT_EXCEL_NAME))
    write_ris(records,   os.path.join(OUTPUT_FOLDER, OUTPUT_RIS_NAME))

    total = len(records)
    n_yes = sum(1 for r in records if r.get("_match") == "YES")
    print("\n" + "=" * 60)
    print("  DONE")
    print("=" * 60)
    print(f"  Total records processed : {total}")
    print(f"  Included (YES)          : {n_yes}")
    print(f"  Excluded (NO)           : {total - n_yes}")
    print(f"\n  Files saved to: {OUTPUT_FOLDER}")
    print(f"    -> {OUTPUT_EXCEL_NAME}")
    print(f"    -> {OUTPUT_RIS_NAME}")
    print("=" * 60)


if __name__ == "__main__":
    main()
